# from locale import DAY_1
from tkinter import *
from tkinter import ttk
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import showinfo
from tkinter import messagebox
from app import generate_invoices
# import os
# import sys
# import pyxlsb
from openpyxl import load_workbook
import calendar

# path function for referencing template
# def resource_path(relative_path):
#     """ Get absolute path to resource, works for dev and for PyInstaller """
#     base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
#     return os.path.join(base_path, relative_path)

# template = resource_path('Invoice Template.docx')

# test funcntion showing file has been selected
def file_select():
    window.filename = askopenfilename()

    if window.filename[-5:] not in [".xlsx",".xlsm",".xltx",".xltm"]:
        # print
        messagebox.showerror("File Error", "File type not supported. Please use open fo the following file types: '.xlsx','.xlsm','.xltx','.xltm'")
    elif window.filename:
        Label(window ,text = window.filename ).grid(row = 0,column = 2)
        sheet_names = load_workbook(window.filename, read_only=True, keep_links=False).sheetnames
        d1 = OptionMenu(window, sheet_name,*sheet_names).grid(row = 3,column = 1)
        sheet_name.set(sheet_names[0])

        print ("selected:", window.filename)
    else:
        print( "file not selected")

# Submit button functionality 
def submit():
    print(year.get())
    print(month.get())
    print(window.filename)

    generate_invoices(invoice_month_num=month_dict[month.get()],invoice_year=year.get(),filename=window.filename,sheet_name=sheet_name.get())# ,template_file=template)
    print("finished generating invoices")
    showinfo("Window", "Invoices Generated!")

# GUI creation
window = Tk()
window.title("Invoice Generator")
window.geometry('500x200')

a = Label(window ,text = "Select File").grid(row = 0,column = 0)
b = Label(window ,text = "Month Number").grid(row = 1,column = 0)
c = Label(window ,text = "Year").grid(row = 2,column = 0)
d = Label(window ,text = "Sheet Name").grid(row=3, column = 0)

year = StringVar(window)
year.set(2021) # default value
month = StringVar(window)
month.set("Jan") # default value
sheet_name = StringVar(window)
# sheet_name.set()
month_dict = {month: index for index, month in enumerate(calendar.month_abbr) if month}
months = [key for key in month_dict.keys()]
a1 = Button(window,text='Browse',command=file_select).grid(row = 0,column = 1)
b1 = OptionMenu(window, month, *months).grid(row = 1,column = 1)
c1 = OptionMenu(window, year, 2020, 2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030).grid(row = 2,column = 1)
# d1 = ''

btn = ttk.Button(window ,text="Submit",command=submit).grid(row=4,column=0)

window.mainloop()